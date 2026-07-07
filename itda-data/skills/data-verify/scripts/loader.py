"""값 grid 로드 — openpyxl 계산값(data_only) / CSV. #967 data-verify.

백엔드 독립(#968 대비): 이 모듈은 (sheet_name -> Grid) 만 공급하고, 판정은 verifiers 가 한다.
COM/OpenXML 백엔드가 같은 Grid 인터페이스를 채우면 verifiers 를 그대로 재사용할 수 있다.
"""
from __future__ import annotations

import csv
from dataclasses import dataclass, field

try:
    import openpyxl
    from openpyxl.utils.cell import coordinate_to_tuple
except ImportError as exc:  # 명시적 에러(조용한 폴백 금지)
    raise SystemExit("openpyxl 필요 — python3 -m pip install -r requirements.txt") from exc


@dataclass
class Grid:
    name: str
    rows: list = field(default_factory=list)   # list[list] 계산값(헤더 포함)

    def header(self) -> list:
        return self.rows[0] if self.rows else []

    def data_rows(self) -> list:
        return self.rows[1:] if self.rows else []

    def col_index(self, name) -> int | None:
        for i, v in enumerate(self.header()):
            if str(v).strip() == str(name).strip():
                return i
        return None

    def cell_a1(self, ref: str):
        r, c = coordinate_to_tuple(ref)   # 1-based (row, col)
        if 1 <= r <= len(self.rows) and 1 <= c <= len(self.rows[r - 1]):
            return self.rows[r - 1][c - 1]
        return None


def _load_csv(path: str) -> Grid:
    for enc in ("utf-8-sig", "utf-8", "cp949"):
        try:
            with open(path, newline="", encoding=enc) as f:
                return Grid(name="Sheet1", rows=[list(r) for r in csv.reader(f)])
        except UnicodeDecodeError:
            continue
    raise ValueError(f"CSV 인코딩 판별 실패(utf-8/cp949): {path}")


def load_sheets(path: str) -> dict:
    """{sheet_name: Grid}. .xlsx 는 전체 시트(계산값), .csv/.tsv 는 단일 Sheet1."""
    if path.lower().endswith((".csv", ".tsv")):
        return {"Sheet1": _load_csv(path)}
    wb = openpyxl.load_workbook(path, data_only=True)
    out = {}
    for name in wb.sheetnames:
        ws = wb[name]
        out[name] = Grid(name=name, rows=[[c.value for c in row] for row in ws.iter_rows()])
    return out

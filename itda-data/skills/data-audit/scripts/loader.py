"""워크북 로드 — openpyxl 수식면·값면 양면 (#952 audit-xls 이식).

audit 은 '누가 =A1*1.05 를 박았나'(수식 문자열)와 '#REF! 가 떴나'(캐시 계산값) 양쪽이 필요하다.
openpyxl 은 한 번의 load 로 둘 다 주지 않는다 — data_only=False 는 수식, True 는 마지막 저장 시 캐시값.
그래서 두 번 로드해 시트별 formulas/computed/literals 로 가른다.

한계: data_only=True 캐시값은 '파일에 마지막 저장된' 값이다. openpyxl 로만 만들고 Excel 로
저장한 적 없는 파일은 캐시가 비어 computed=None 이 된다(SKILL.md 정확성 한계 절 참조).
"""
from __future__ import annotations

from dataclasses import dataclass

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # 명시적 에러 표면화(조용한 폴백 금지)
    raise SystemExit(
        "openpyxl 이 필요합니다 — python3 -m pip install -r requirements.txt"
    ) from exc

ERROR_CODES = ("#REF!", "#VALUE!", "#N/A", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!")


def a1(row: int, col: int) -> str:
    return f"{get_column_letter(col)}{row}"


def is_error_value(val) -> bool:
    return isinstance(val, str) and val in ERROR_CODES


@dataclass
class SheetView:
    name: str
    state: str                 # 'visible' | 'hidden' | 'veryHidden'
    max_row: int
    max_col: int
    formulas: dict             # (row, col) -> 수식 문자열 '=...'
    computed: dict             # (row, col) -> 수식의 캐시 계산값(에러 문자열 포함, None 가능)
    literals: dict             # (row, col) -> 정적 값(수식 아님)
    hidden_rows: list          # [int, ...]
    hidden_cols: list          # ['A', ...]


def load_views(path: str, sheet: str | None = None):
    """(list[SheetView], list[all_sheet_names]) 반환. sheet=None 이면 전체(숨은 시트 포함)."""
    wb_f = openpyxl.load_workbook(path, data_only=False)
    wb_v = openpyxl.load_workbook(path, data_only=True)
    all_names = list(wb_f.sheetnames)
    if sheet is not None and sheet not in all_names:
        raise ValueError(f"시트를 찾을 수 없습니다: {sheet!r} (가능: {all_names})")
    targets = all_names if sheet is None else [sheet]

    views: list[SheetView] = []
    for name in targets:
        ws_f, ws_v = wb_f[name], wb_v[name]
        max_row = ws_f.max_row or 0
        max_col = ws_f.max_column or 0
        formulas: dict = {}
        computed: dict = {}
        literals: dict = {}
        rows_f = ws_f.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col)
        rows_v = ws_v.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col)
        for row_f, row_v in zip(rows_f, rows_v):
            for cf, cv in zip(row_f, row_v):
                rc = (cf.row, cf.column)
                v = cf.value
                if cf.data_type == "f" or (isinstance(v, str) and v.startswith("=")):
                    formulas[rc] = v
                    computed[rc] = cv.value
                elif v is not None:
                    literals[rc] = v
        hidden_rows = sorted(i for i, d in ws_f.row_dimensions.items() if d.hidden)
        hidden_cols = sorted(k for k, d in ws_f.column_dimensions.items() if d.hidden)
        views.append(
            SheetView(
                name=name,
                state=ws_f.sheet_state,
                max_row=max_row,
                max_col=max_col,
                formulas=formulas,
                computed=computed,
                literals=literals,
                hidden_rows=hidden_rows,
                hidden_cols=hidden_cols,
            )
        )
    return views, all_names

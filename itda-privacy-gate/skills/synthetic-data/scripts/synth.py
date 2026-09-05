#!/usr/bin/env python3
"""synthetic-data — 프리셋/스펙(JSON) 으로 가상 데이터를 만들고 규칙을 검증한다.

    python3 synth.py presets                                   # 프리셋 목록
    python3 synth.py show   <도메인/문서>                        # 프리셋 JSON 출력(사용자가 고칠 출발점)
    python3 synth.py validate <스펙.json | 도메인/문서>           # 스키마·규칙 참조·등급 검증 (RED = exit 1)
    python3 synth.py generate <스펙.json | 도메인/문서> --rows N --out DIR --confirm-fake
                     [--seed S] [--xlsx-template 양식.xlsx] [--hwpx-template 양식.hwpx] [--hwpx-rows 3]
    python3 synth.py render <DIR/data.json> [--xlsx-template …] [--hwpx-template …]   # fill-text 뒤 재렌더(행 재생성 없음)
    python3 synth.py verify <스펙.json> <DIR/data.json>          # 규칙 재검증 리포트만
    python3 synth.py fill-text <DIR/data.json> <texts.json>     # 자유텍스트 placeholder 를 AI 문장으로 치환

산출(DIR): data.json · data.csv · <문서>.xlsx · report.md(검증 리포트+한계 고지) · field-definitions.md(항목 정의표)
exit: 0 성공 · 1 검증 RED(규칙 위반·스키마 위반) · 2 사용 오류
표·규칙·검증은 전부 이 코드가 결정론으로 한다. LLM 은 자유텍스트 문장만 쓴다(fill-text).
"""
from __future__ import annotations

import argparse
import csv
import datetime as dt
import hashlib
import json
import random
import re
import sys
import zipfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
import identifiers as ident  # noqa: E402

SKILL_DIR = Path(__file__).resolve().parent.parent
PRESETS_DIR = SKILL_DIR / "presets"

GRADES = ("식별자", "민감정보", "준식별자", "비개인정보")
TYPES = ("string", "int", "float", "date", "text")
GEN_KINDS = ("sequence", "person_name", "rrn", "phone", "address", "choice", "int", "float",
             "date", "date_after", "derive", "code", "free_text", "const", "time_slot")
RULE_KINDS = ("compare", "derived", "prefix_map", "unique", "in_set", "not_empty", "empty_iff")
DERIVE_FNS = ("days_between", "prefix_map", "age_from", "sum", "diff")
LEGAL_STATUS = ("확인", "확인 필요")

PRESET_NOTICE = ("이 프리셋의 항목·규칙은 해당 업무의 문서 구조를 예측해 만든 것일 뿐 정확하지 않습니다. "
                 "실제 업무의 데이터 구조(항목 이름·규칙·서식)를 알려주시면 그에 맞춰 생성할 수 있습니다.")
DATA_NOTICE = ("이 데이터는 전부 가상입니다. 실제 환자·직원·거래처가 아니며, 항목별 분포만 닮았고 항목 간 상관관계는 없습니다. "
               "스크립트·수식·절차 검증용이며 경영 판단 수치로 쓰지 않습니다. "
               "이름은 가상 성명(동명이인 의도 삽입), 주민번호는 검증자리를 일부러 틀린 번호, 연락처는 미배정 대역, 주소는 가상 지명입니다.")
UNMODIFIED_NOTICE = "프리셋 그대로 생성 — 본인 서식과 다를 수 있음."
FREE_TEXT_PLACEHOLDER = "«자유텍스트 미작성: {field} — AI 가 가상 문장을 씁니다(fill-text)»"


class SpecError(Exception):
    pass


CONFIRM_FLAG = "--confirm-fake"
CONFIRM_TEXT = "이 예시는 실제 환자·직원이 아닌 것을 확인합니다."
_RRN_RE = re.compile(r"\b\d{6}-?\d{7}\b")
_PHONE_RE = re.compile(r"\b01[016789]-?\d{3,4}-?\d{4}\b")


def pii_findings(text: str) -> list[str]:
    """실제 데이터 유입 차단 — 검증식을 통과하는 주민번호·부여 가능 대역의 휴대전화가 보이면 지목한다.
    우리 생성기가 만드는 값(검증자리 불일치·010-0000)은 통과한다."""
    out = []
    for m in _RRN_RE.finditer(text):
        if ident.rrn_is_valid(m.group(0)):
            out.append(f"검증식 통과 주민번호 형태: {m.group(0)[:8]}…")
    for m in _PHONE_RE.finditer(text):
        digits = re.sub(r"\D", "", m.group(0))
        if not digits.startswith("0100000"):
            out.append(f"부여 가능 대역 휴대전화: {m.group(0)[:7]}…")
    return out


def scan_pii(values) -> list[str]:
    found = []
    for v in values:
        if isinstance(v, str):
            found += pii_findings(v)
    return found


# ---------------------------------------------------------------- presets / spec

def list_presets() -> list[str]:
    out = []
    for p in sorted(PRESETS_DIR.glob("*/*.json")):
        out.append(f"{p.parent.name}/{p.stem}")
    return out


def resolve_spec_path(ref: str) -> Path:
    p = Path(ref)
    if p.suffix == ".json" and p.exists():
        return p
    cand = PRESETS_DIR / f"{ref}.json"
    if cand.exists():
        return cand
    raise SpecError(f"스펙/프리셋을 찾을 수 없습니다: {ref} (목록: python3 synth.py presets)")


def load_spec(ref: str) -> dict:
    return json.loads(resolve_spec_path(ref).read_text(encoding="utf-8"))


def structure_hash(spec: dict) -> str:
    core = {"fields": spec.get("fields"), "rules": spec.get("rules")}
    return hashlib.sha256(json.dumps(core, ensure_ascii=False, sort_keys=True).encode()).hexdigest()


def is_unmodified_preset(spec: dict) -> bool:
    """스펙의 fields·rules 가 같은 domain/document 프리셋과 동일하면 '프리셋 그대로'."""
    cand = PRESETS_DIR / str(spec.get("domain", "")) / f"{spec.get('document', '')}.json"
    if not cand.exists():
        return False
    return structure_hash(json.loads(cand.read_text(encoding="utf-8"))) == structure_hash(spec)


# ---------------------------------------------------------------- validate

def validate_spec(spec: dict) -> list[str]:
    errs: list[str] = []
    for k in ("schema_version", "domain", "document", "title", "format", "fields", "rules", "reid_keys", "legal", "preset_notice"):
        if k not in spec:
            errs.append(f"[필수키] '{k}' 없음")
    if errs:
        return errs
    if str(spec["schema_version"]) != "1.0":
        errs.append(f"[schema_version] '1.0' 만 지원: {spec['schema_version']}")
    if spec["format"] not in ("xlsx", "hwpx", "docx"):
        errs.append(f"[format] 'xlsx|hwpx|docx' 만 허용: {spec['format']}")
    if spec["preset_notice"] != PRESET_NOTICE:
        errs.append("[preset_notice] 정본 문구와 다름(코드의 PRESET_NOTICE 와 바이트 동일해야 한다)")
    fields = spec["fields"]
    if not isinstance(fields, list) or not fields:
        return errs + ["[fields] 비어 있음"]
    names: list[str] = []
    for i, f in enumerate(fields):
        nm = f.get("name")
        if not nm:
            errs.append(f"[fields[{i}]] name 없음"); continue
        if nm in names:
            errs.append(f"[fields] 항목명 중복: {nm}")
        names.append(nm)
        if f.get("type") not in TYPES:
            errs.append(f"[{nm}] type 은 {TYPES} 중 하나: {f.get('type')}")
        if f.get("grade") not in GRADES:
            errs.append(f"[{nm}] 개인정보 등급 누락/오류(허용: {'/'.join(GRADES)}): {f.get('grade')}")
        g = f.get("generator") or {}
        kind = g.get("kind")
        if kind not in GEN_KINDS:
            errs.append(f"[{nm}] generator.kind 미지: {kind}")
        if f.get("free_text") and kind != "free_text":
            errs.append(f"[{nm}] free_text 항목의 generator 는 free_text 여야 함")
        if kind == "free_text" and not f.get("free_text"):
            errs.append(f"[{nm}] generator free_text 인데 free_text:true 가 없음")
        if kind == "choice" and not g.get("values"):
            errs.append(f"[{nm}] choice.values 비어 있음")
        if kind in ("date_after",) and not g.get("from"):
            errs.append(f"[{nm}] date_after.from 없음")
        if kind == "derive":
            if g.get("fn") not in DERIVE_FNS:
                errs.append(f"[{nm}] derive.fn 미지: {g.get('fn')}")
            if not g.get("args"):
                errs.append(f"[{nm}] derive.args 없음")
            if g.get("fn") == "prefix_map" and not isinstance(g.get("map"), dict):
                errs.append(f"[{nm}] derive prefix_map 은 map(dict) 필요")
        if kind == "rrn" and g.get("birth_from") and g["birth_from"] not in names + [x.get("name") for x in fields]:
            errs.append(f"[{nm}] rrn.birth_from 참조 오류: {g['birth_from']}")
    all_names = set(names)
    # generator 참조
    for f in fields:
        g = f.get("generator") or {}
        for ref in [g.get(k) for k in ("from", "birth_from", "gender_from", "empty_when_empty") if g.get(k)] + list(g.get("args") or []):
            if isinstance(ref, str) and ref not in all_names:
                errs.append(f"[{f.get('name')}] generator 참조 오류: '{ref}' 항목 없음")
    # rules
    rules = spec["rules"]
    if not isinstance(rules, list):
        errs.append("[rules] 리스트여야 함")
        rules = []
    for i, r in enumerate(rules):
        kind = r.get("kind")
        if kind not in RULE_KINDS:
            errs.append(f"[rules[{i}]] kind 미지: {kind}"); continue
        if not r.get("id"):
            errs.append(f"[rules[{i}]] id 없음")
        refs = []
        if kind == "compare":
            refs = [r.get("left"), r.get("right")]
            if r.get("op") not in ("<", "<=", "==", ">=", ">", "!="):
                errs.append(f"[rules[{i}]] compare.op 오류: {r.get('op')}")
        elif kind == "derived":
            refs = [r.get("field")] + list(r.get("args") or [])
            if r.get("fn") not in DERIVE_FNS:
                errs.append(f"[rules[{i}]] derived.fn 미지: {r.get('fn')}")
        elif kind == "prefix_map":
            refs = [r.get("field"), r.get("from")]
            if not isinstance(r.get("map"), dict):
                errs.append(f"[rules[{i}]] prefix_map.map 필요")
        elif kind in ("unique", "not_empty"):
            refs = [r.get("field")]
        elif kind == "empty_iff":
            refs = [r.get("field"), r.get("with")]
        elif kind == "in_set":
            refs = [r.get("field")]
            if not r.get("values"):
                errs.append(f"[rules[{i}]] in_set.values 비어 있음")
        for ref in refs:
            if ref not in all_names:
                errs.append(f"[rules[{i}]:{r.get('id')}] 참조 오류: '{ref}' 항목 없음")
    for k in spec.get("reid_keys") or []:
        if k not in all_names:
            errs.append(f"[reid_keys] 참조 오류: '{k}' 항목 없음")
    for j, L in enumerate(spec.get("legal") or []):
        if L.get("status") not in LEGAL_STATUS:
            errs.append(f"[legal[{j}]] status 는 {LEGAL_STATUS} 중 하나")
        if not L.get("law") or not L.get("article"):
            errs.append(f"[legal[{j}]] law/article 필요")
    free_names = {f.get("name") for f in fields if f.get("free_text")}
    for i, row in enumerate(spec.get("seed_examples") or []):
        for k, v in row.items():
            if k not in all_names:
                errs.append(f"[seed_examples[{i}]] 항목명 오류: {k}")
            if k in free_names and v not in ("", None):
                errs.append(f"[seed_examples[{i}]] 자유텍스트 항목 '{k}' 에는 예시를 받지 않는다(철칙 2)")
        for hit in scan_pii(row.values()):
            errs.append(f"[seed_examples[{i}]] 실제 데이터 의심 — {hit}")
    return errs


# ---------------------------------------------------------------- generate

def _order_fields(fields: list[dict]) -> list[dict]:
    """generator 참조(from/args/birth_from/gender_from) 를 먼저 만들도록 위상 정렬."""
    by = {f["name"]: f for f in fields}
    done: list[str] = []
    seen: set[str] = set()

    def deps(f):
        g = f.get("generator") or {}
        d = []
        for k in ("from", "birth_from", "gender_from", "empty_when_empty"):
            if g.get(k):
                d.append(g[k])
        d += [a for a in (g.get("args") or []) if isinstance(a, str) and a in by]
        return d

    def visit(n, stack=()):
        if n in seen:
            return
        if n in stack:
            raise SpecError(f"generator 참조 순환: {' -> '.join(stack + (n,))}")
        for d in deps(by[n]):
            visit(d, stack + (n,))
        seen.add(n); done.append(n)

    for f in fields:
        visit(f["name"])
    return [by[n] for n in done]


def _pdate(v):
    if v in (None, ""):
        return None
    return dt.date.fromisoformat(str(v)[:10])


def _num(v):
    f = float(v)
    return int(f) if f.is_integer() and not isinstance(v, float) else f


def _derive(fn: str, args: list, g: dict, row: dict):
    if fn == "days_between":
        a, b = _pdate(row.get(args[0])), _pdate(row.get(args[1]))
        return "" if a is None or b is None else (b - a).days
    if fn == "diff":
        a, b = row.get(args[0]), row.get(args[1])
        return "" if a in ("", None) or b in ("", None) else _num(a) - _num(b)
    if fn == "sum":
        vals = [row.get(a) for a in args]
        return "" if any(v in ("", None) for v in vals) else sum(_num(v) for v in vals)
    if fn == "prefix_map":
        src = str(row.get(args[0], ""))
        n = int(g.get("length", 1))
        return g["map"].get(src[:n], "")
    if fn == "age_from":
        b = _pdate(row.get(args[0]))
        ref = _pdate(row.get(args[1])) if len(args) > 1 else dt.date(2026, 1, 1)
        if b is None or ref is None:
            return ""
        return ref.year - b.year - ((ref.month, ref.day) < (b.month, b.day))
    raise SpecError(f"derive fn 미지: {fn}")


def generate_rows(spec: dict, n: int, seed: int = 7) -> list[dict]:
    rng = random.Random(seed)
    fields = _order_fields(spec["fields"])
    rows = [dict() for _ in range(n)]
    seeds = spec.get("seed_examples") or []
    for f in fields:
        nm, g = f["name"], f.get("generator") or {}
        kind = g["kind"]
        if kind == "person_name":
            vals = ident.person_names(rng, n, float(g.get("dup_ratio", 0.06)))
            for i in range(n):
                rows[i][nm] = vals[i]
            continue
        for i in range(n):
            row = rows[i]
            if kind == "sequence":
                row[nm] = f"{g.get('prefix', '')}{int(g.get('start', 1)) + i:0{int(g.get('width', 4))}d}"
            elif kind == "rrn":
                b = _pdate(row.get(g["birth_from"])) if g.get("birth_from") else None
                gd = row.get(g["gender_from"]) if g.get("gender_from") else None
                row[nm] = ident.rrn(rng, b, gd)
            elif kind == "phone":
                row[nm] = ident.phone(rng)
            elif kind == "address":
                row[nm] = ident.address(rng)
            elif kind == "choice":
                if g.get("empty_when_empty") and row.get(g["empty_when_empty"]) in ("", None):
                    row[nm] = ""; continue
                vals = list(g["values"])
                # 사용자 가짜 예시의 값을 후보에 편입(구조는 진짜, 값은 가짜)
                for s in seeds:
                    v = s.get(nm)
                    if v not in (None, "") and v not in vals:
                        vals.append(v)
                w = g.get("weights")
                if w and len(w) == len(vals):
                    row[nm] = rng.choices(vals, weights=w)[0]
                else:
                    row[nm] = rng.choice(vals)
            elif kind == "int":
                row[nm] = rng.randint(int(g.get("min", 0)), int(g.get("max", 100)))
            elif kind == "float":
                row[nm] = round(rng.uniform(float(g.get("min", 0)), float(g.get("max", 100))), int(g.get("digits", 1)))
            elif kind == "date":
                s, e = _pdate(g.get("start", "2025-01-01")), _pdate(g.get("end", "2025-12-31"))
                row[nm] = (s + dt.timedelta(days=rng.randrange(0, (e - s).days + 1))).isoformat()
            elif kind == "date_after":
                base = _pdate(row.get(g["from"]))
                if base is None or rng.random() < float(g.get("null_ratio", 0.0)):
                    row[nm] = ""
                else:
                    row[nm] = (base + dt.timedelta(days=rng.randint(int(g.get("min_days", 0)), int(g.get("max_days", 30))))).isoformat()
            elif kind == "derive":
                row[nm] = _derive(g["fn"], g["args"], g, row)
            elif kind == "code":
                row[nm] = ident.code(rng, g.get("pattern", "A####"))
            elif kind == "free_text":
                row[nm] = FREE_TEXT_PLACEHOLDER.format(field=nm)
            elif kind == "const":
                row[nm] = g.get("value", "")
            elif kind == "time_slot":
                slots = g.get("values") or ["D", "E", "N", "OFF"]
                row[nm] = rng.choice(slots)
            else:
                raise SpecError(f"generator kind 미지: {kind}")
    order = [f["name"] for f in spec["fields"]]
    return [{k: r.get(k, "") for k in order} for r in rows]


# ---------------------------------------------------------------- verify

def _cmp(a, b, op) -> bool:
    return {"<": a < b, "<=": a <= b, "==": a == b, ">=": a >= b, ">": a > b, "!=": a != b}[op]


def _coerce(a, b):
    da, db = None, None
    try:
        da, db = _pdate(a), _pdate(b)
    except ValueError:
        pass
    if da is not None and db is not None:
        return da, db
    try:
        return float(a), float(b)
    except (TypeError, ValueError):
        return str(a), str(b)


def verify_rows(spec: dict, rows: list[dict]) -> dict:
    results = []
    for r in spec["rules"]:
        viol = []
        kind = r["kind"]
        seen = {}
        for i, row in enumerate(rows):
            if kind == "compare":
                a, b = row.get(r["left"]), row.get(r["right"])
                if a in ("", None) or b in ("", None):
                    continue  # 미기재(재원 중 등)는 비교 대상 아님
                ca, cb = _coerce(a, b)
                if not _cmp(ca, cb, r["op"]):
                    viol.append(i + 1)
            elif kind == "derived":
                exp = _derive(r["fn"], r["args"], r, row)
                got = row.get(r["field"])
                if str(exp) != str(got) and not (exp == "" and got in ("", None)):
                    try:
                        if float(exp) == float(got):
                            continue
                    except (TypeError, ValueError):
                        pass
                    viol.append(i + 1)
            elif kind == "prefix_map":
                src = str(row.get(r["from"], ""))
                exp = r["map"].get(src[:int(r.get("length", 1))])
                if exp is None or str(row.get(r["field"])) != str(exp):
                    viol.append(i + 1)
            elif kind == "unique":
                v = row.get(r["field"])
                if v in seen:
                    viol.append(i + 1)
                seen[v] = True
            elif kind == "in_set":
                if row.get(r["field"]) not in r["values"]:
                    viol.append(i + 1)
            elif kind == "not_empty":
                if row.get(r["field"]) in ("", None):
                    viol.append(i + 1)
            elif kind == "empty_iff":
                if (row.get(r["field"]) in ("", None)) != (row.get(r["with"]) in ("", None)):
                    viol.append(i + 1)
        results.append({"id": r["id"], "kind": kind, "desc": r.get("desc", ""), "violations": len(viol), "rows": viol[:20]})
    placeholders = sum(1 for row in rows for v in row.values() if isinstance(v, str) and v.startswith("«자유텍스트 미작성"))
    return {"rows": len(rows), "rules": results, "total_violations": sum(x["violations"] for x in results),
            "free_text_placeholders": placeholders}


# ---------------------------------------------------------------- outputs

def write_csv(path: Path, rows: list[dict], fields: list[str]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as fp:
        w = csv.DictWriter(fp, fieldnames=fields)
        w.writeheader(); w.writerows(rows)


def field_definitions_md(spec: dict) -> str:
    lines = [f"# 항목 정의표 — {spec['title']} ({spec['domain']}/{spec['document']})", "",
             "1교시 「판단 기준표」에 붙이는 형태. 등급은 프리셋의 예측이며 기관 내부 기준으로 고쳐 쓴다.", "",
             "| 항목 | 타입 | 개인정보 등급 | 자유텍스트 | 생성 방식 | 설명 |", "|---|---|---|---|---|---|"]
    for f in spec["fields"]:
        g = f.get("generator") or {}
        lines.append(f"| {f['name']} | {f['type']} | **{f['grade']}** | {'예' if f.get('free_text') else ''} | {g.get('kind')} | {f.get('desc', '')} |")
    lines += ["", f"재식별 게임용 준식별자 조합(reid_keys): {', '.join(spec.get('reid_keys') or []) or '(없음)'}", ""]
    if spec.get("legal"):
        lines += ["## 근거 조문", ""]
        for L in spec["legal"]:
            lines.append(f"- {L['law']} {L['article']} — {L.get('note', '')} ({L['status']})")
    return "\n".join(lines) + "\n"


def report_md(spec: dict, res: dict, unmodified: bool, seed: int, extra: list[str]) -> str:
    lines = [f"# 가상 데이터 생성 리포트 — {spec['title']}", "", "## 한계 고지", "",
             f"> **데이터 한계** — {DATA_NOTICE}", "",
             f"> **프리셋 한계** — {spec['preset_notice']}", ""]
    if unmodified:
        lines += [f"> ⚠️ **{UNMODIFIED_NOTICE}**", ""]
    lines += ["## 규칙 검증", "", f"- 행 수: {res['rows']} · seed: {seed}", f"- 규칙 {len(res['rules'])}개 · 위반 합계 **{res['total_violations']}**", "",
              "| 규칙 id | 종류 | 설명 | 위반 |", "|---|---|---|---|"]
    for r in res["rules"]:
        rows = f" (행 {', '.join(map(str, r['rows']))})" if r["rows"] else ""
        lines.append(f"| {r['id']} | {r['kind']} | {r['desc']} | {r['violations']}{rows} |")
    lines += ["", f"- 자유텍스트 미작성 placeholder: {res['free_text_placeholders']}건" + ("  ← `fill-text` 로 AI 가상 문장을 채운다" if res["free_text_placeholders"] else ""), ""]
    if extra:
        lines += ["## 산출 파일", ""] + [f"- {x}" for x in extra] + [""]
    return "\n".join(lines)


# ---------------------------------------------------------------- xlsx

def write_xlsx(path: Path, spec: dict, rows: list[dict], template: Path | None, notice_lines: list[str]) -> str:
    try:
        import openpyxl  # noqa
    except ImportError:
        return "openpyxl 미설치 — xlsx 생략(pip install openpyxl)"
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter
    fields = [f["name"] for f in spec["fields"]]
    if template:
        wb = load_workbook(template)
        ws, header_row, col_of = _locate_header(wb, fields)
        if ws is None:
            raise SpecError(f"양식에서 항목명 헤더 행을 찾지 못했습니다(항목의 60% 이상이 한 행에 있어야 함): {template}")
        style_src = header_row + 1 if ws.max_row > header_row else None
        occupied = sum(1 for r in range(header_row + 1, min(ws.max_row, header_row + len(rows)) + 1)
                       for c in col_of.values() if ws.cell(row=r, column=c).value not in (None, ""))
        shifted = ""
        if occupied:
            ws.insert_rows(header_row + 1, amount=len(rows))
            style_src = header_row + 1 + len(rows)
            shifted = f" · 기존 내용 {occupied}셀을 {len(rows)}행 아래로 밀었다(수식 참조는 갱신되지 않는다 — 빈 양식 권장)"
        for i, row in enumerate(rows):
            r = header_row + 1 + i
            for nm, c in col_of.items():
                cell = ws.cell(row=r, column=c, value=row.get(nm, ""))
                if style_src and style_src != r:
                    src = ws.cell(row=style_src, column=c)
                    if src.has_style:
                        cell._style = src._style
        info = f"양식 '{ws.title}' 시트 {header_row}행 헤더 아래 {len(rows)}행 기입(매칭 항목 {len(col_of)}/{len(fields)}){shifted}"
    else:
        wb = Workbook(); ws = wb.active; ws.title = spec["title"][:31]
        ws.append(fields)
        for row in rows:
            ws.append([row.get(k, "") for k in fields])
        for i, nm in enumerate(fields, 1):
            ws.column_dimensions[get_column_letter(i)].width = max(10, min(40, len(str(nm)) * 2 + 4))
        info = f"새 통합문서 '{ws.title}' 시트 {len(rows)}행"
    ns = wb.create_sheet("안내", 0)
    for line in notice_lines:
        ns.append([line])
    ns.column_dimensions["A"].width = 120
    wb.save(path)
    return info


def _locate_header(wb, fields: list[str]):
    best = (None, 0, {}, 0)
    for ws in wb.worksheets:
        for r in range(1, min(ws.max_row, 30) + 1):
            col_of = {}
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                if isinstance(v, str) and v.strip() in fields:
                    col_of.setdefault(v.strip(), c)
            if len(col_of) > best[3]:
                best = (ws, r, col_of, len(col_of))
    ws, r, col_of, n = best
    if ws is None or n < max(1, int(len(fields) * 0.6)):
        return None, 0, {}
    return ws, r, col_of


# ---------------------------------------------------------------- hwpx (양식 1건 = 1행)

_SECTION_RE = re.compile(r"^Contents/section\d+\.xml$")
_TEXT_RE = re.compile(r"(<hp:t[^>]*>)([^<]*)(</hp:t>)")
_RUN_MERGE_RE = re.compile(r"<hp:run([^>]*)><hp:t>([^<]*)</hp:t></hp:run><hp:run\1><hp:t>([^<]*)</hp:t></hp:run>")


def _xml_escape(s: str) -> str:
    return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _fill_xml(xml: str, mapping: dict[str, str]) -> tuple[str, dict[str, int]]:
    prev = None
    while prev != xml:
        prev = xml
        xml = _RUN_MERGE_RE.sub(lambda m: f"<hp:run{m.group(1)}><hp:t>{m.group(2)}{m.group(3)}</hp:t></hp:run>", xml)
    counts = {k: 0 for k in mapping}

    def rep(m):
        text = m.group(2)
        for k, v in mapping.items():
            ke = _xml_escape(k)
            if ke in text:
                counts[k] += text.count(ke)
                text = text.replace(ke, _xml_escape(v))
        return m.group(1) + text + m.group(3)
    return _TEXT_RE.sub(rep, xml), counts


def fill_hwpx(template: Path, out: Path, mapping: dict[str, str]) -> dict[str, int]:
    """(항목명)·{{항목명}} placeholder 를 값으로 치환. mimetype 첫 엔트리·STORED 보존. 원본은 건드리지 않는다."""
    total = {k: 0 for k in mapping}
    with zipfile.ZipFile(template) as zin, zipfile.ZipFile(out, "w") as zout:
        names = zin.namelist()
        if "mimetype" in names:
            zout.writestr(zipfile.ZipInfo("mimetype"), zin.read("mimetype"), compress_type=zipfile.ZIP_STORED)
        for n in names:
            if n == "mimetype":
                continue
            data = zin.read(n)
            if _SECTION_RE.match(n):
                xml, counts = _fill_xml(data.decode("utf-8"), mapping)
                for k, c in counts.items():
                    total[k] += c
                data = xml.encode("utf-8")
            zout.writestr(n, data, compress_type=zipfile.ZIP_DEFLATED)
    return total


def hwpx_mapping(row: dict) -> dict[str, str]:
    m = {}
    for k, v in row.items():
        m[f"({k})"] = str(v); m[f"{{{{{k}}}}}"] = str(v)
    return m


# ---------------------------------------------------------------- commands

def render_outputs(spec: dict, rows: list[dict], seed: int, out: Path, unmodified: bool,
                   xlsx_template: str | None, hwpx_template: str | None, hwpx_rows: int) -> tuple[dict, list[str]]:
    """data.json 은 호출자가 이미 썼다고 가정. csv·정의표·xlsx·hwpx·report 를 만든다. 원본 양식은 읽기만."""
    fields = [f["name"] for f in spec["fields"]]
    res = verify_rows(spec, rows)
    write_csv(out / "data.csv", rows, fields)
    (out / "field-definitions.md").write_text(field_definitions_md(spec), encoding="utf-8")
    extra = ["data.json", "data.csv", "field-definitions.md", "report.md"]
    notice = ["가상 데이터 — 한계 고지", DATA_NOTICE, spec["preset_notice"]] + ([UNMODIFIED_NOTICE] if unmodified else [])
    xlsx_path = out / f"{spec['document']}.xlsx"
    hwpx_paths = [out / f"{spec['document']}-{i:03d}.hwpx" for i in range(1, min(hwpx_rows, len(rows)) + 1)] if hwpx_template else []
    # 원본 양식 보호 — 쓰기 전에 출력 경로 전부를 확정하고 충돌을 일괄 검사한다(P1-1)
    for tpl in (xlsx_template, hwpx_template):
        if tpl and Path(tpl).resolve() in {q.resolve() for q in [xlsx_path, *hwpx_paths]}:
            raise SpecError(f"원본 양식을 덮어쓰지 않습니다 — --out 을 다른 디렉토리로 주세요: {tpl}")
    if xlsx_template and not Path(xlsx_template).exists():
        raise SpecError(f"xlsx 양식 없음: {xlsx_template}")
    if hwpx_template and not Path(hwpx_template).exists():
        raise SpecError(f"hwpx 양식 없음: {hwpx_template}")
    info = write_xlsx(xlsx_path, spec, rows, Path(xlsx_template) if xlsx_template else None, notice)
    extra.append(f"{xlsx_path.name} — {info}")
    for i, dst in enumerate(hwpx_paths):
        mapping = hwpx_mapping(rows[i])
        notice_text = " ".join(notice[1:])
        mapping["(한계고지)"] = notice_text; mapping["{{한계고지}}"] = notice_text
        tmp = dst.with_suffix(".hwpx.part")
        counts = fill_hwpx(Path(hwpx_template), tmp, mapping)
        tmp.replace(dst)
        hit = sum(1 for k, c in counts.items() if c and k.startswith("(") and k != "(한계고지)")
        notice_hit = counts["(한계고지)"] + counts["{{한계고지}}"]
        line = f"{dst.name} — placeholder 치환 {hit}/{len(rows[i])} 항목"
        if hit == 0:
            line += " ⚠️ 0건: 양식에 (항목명) placeholder 가 없음"
        if notice_hit == 0:
            line += " ⚠️ 한계 고지 미기입: 양식에 (한계고지) placeholder 가 없어 hwpx 안에는 고지가 없다 — report.md 와 함께 전달"
        extra.append(line)
    (out / "report.md").write_text(report_md(spec, res, unmodified, seed, extra), encoding="utf-8")
    return res, extra


def cmd_generate(a) -> int:
    if not a.confirm_fake:
        print(f"오류: 첫 질문 미확인 — 사용자가 「{CONFIRM_TEXT}」 를 확인한 뒤 {CONFIRM_FLAG} 를 붙여 실행한다(철칙 1)", file=sys.stderr)
        return 2
    if a.rows < 1:
        print("오류: --rows 는 1 이상", file=sys.stderr); return 2
    spec = load_spec(a.spec)
    errs = validate_spec(spec)
    if errs:
        print("스펙 검증 RED:\n  " + "\n  ".join(errs), file=sys.stderr); return 1
    out = Path(a.out); out.mkdir(parents=True, exist_ok=True)
    rows = generate_rows(spec, a.rows, a.seed)
    unmodified = is_unmodified_preset(spec) or resolve_spec_path(a.spec).resolve().is_relative_to(PRESETS_DIR.resolve())
    (out / "data.json").write_text(json.dumps({"spec": spec, "seed": a.seed, "unmodified": unmodified, "rows": rows}, ensure_ascii=False, indent=1), encoding="utf-8")
    res, extra = render_outputs(spec, rows, a.seed, out, unmodified, a.xlsx_template, a.hwpx_template, a.hwpx_rows)
    print(f"생성 {len(rows)}행 → {out}  (규칙 위반 {res['total_violations']} · 자유텍스트 placeholder {res['free_text_placeholders']})")
    for x in extra:
        print("  - " + x)
    if unmodified:
        print("  ⚠️ " + UNMODIFIED_NOTICE)
    return 0 if res["total_violations"] == 0 else 1


def cmd_render(a) -> int:
    """data.json(fill-text 반영본)으로 csv·xlsx·hwpx·report 를 다시 만든다 — 행 재생성 없음."""
    p = Path(a.data)
    data = json.loads(p.read_text(encoding="utf-8"))
    spec, rows = data["spec"], data["rows"]
    if not rows:
        print("오류: 행이 0건", file=sys.stderr); return 2
    res, extra = render_outputs(spec, rows, data.get("seed", -1), p.parent, bool(data.get("unmodified")), a.xlsx_template, a.hwpx_template, a.hwpx_rows)
    print(f"재렌더 {len(rows)}행 → {p.parent}  (규칙 위반 {res['total_violations']} · 자유텍스트 placeholder {res['free_text_placeholders']})")
    for x in extra:
        print("  - " + x)
    return 0 if res["total_violations"] == 0 else 1


def cmd_validate(a) -> int:
    spec = load_spec(a.spec)
    errs = validate_spec(spec)
    if errs:
        print("RED:\n  " + "\n  ".join(errs)); return 1
    print(f"OK: {spec['domain']}/{spec['document']} — 항목 {len(spec['fields'])} · 규칙 {len(spec['rules'])}")
    return 0


def cmd_verify(a) -> int:
    spec = load_spec(a.spec)
    data = json.loads(Path(a.data).read_text(encoding="utf-8"))
    rows = data["rows"] if isinstance(data, dict) else data
    if not rows:
        print("RED: 행이 0건 — 빈 데이터는 규칙을 공허하게 통과하므로 검증 성공으로 치지 않는다"); return 1
    res = verify_rows(spec, rows)
    print(report_md(spec, res, is_unmodified_preset(spec), data.get("seed", -1) if isinstance(data, dict) else -1, []))
    return 0 if res["total_violations"] == 0 else 1


def cmd_fill_text(a) -> int:
    p = Path(a.data)
    data = json.loads(p.read_text(encoding="utf-8"))
    texts = json.loads(Path(a.texts).read_text(encoding="utf-8"))  # {"항목명": {"1": "...", "2": "..."}}
    spec, rows = data["spec"], data["rows"]
    free = {f["name"] for f in spec["fields"] if f.get("free_text")}
    n = 0
    for field, m in texts.items():
        if field not in free:
            print(f"오류: '{field}' 는 자유텍스트 항목이 아닙니다(자유텍스트: {sorted(free)})", file=sys.stderr); return 2
        hits = scan_pii(m.values())
        if hits:
            print("오류: 자유텍스트에 실제 데이터 의심 값 — 실제 문장은 예시로도 받지 않는다(철칙 2): " + "; ".join(hits), file=sys.stderr); return 2
        for k, v in m.items():
            i = int(k) - 1
            if 0 <= i < len(rows):
                rows[i][field] = v; n += 1
    p.write_text(json.dumps(data, ensure_ascii=False, indent=1), encoding="utf-8")
    write_csv(p.parent / "data.csv", rows, [f["name"] for f in spec["fields"]])
    left = verify_rows(spec, rows)["free_text_placeholders"]
    print(f"자유텍스트 {n}건 반영 · 남은 placeholder {left}건 → `render {p}` 로 xlsx·hwpx·report 를 다시 만든다")
    return 0


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    sub = ap.add_subparsers(dest="cmd", required=True)
    sub.add_parser("presets")
    s = sub.add_parser("show"); s.add_argument("spec")
    s = sub.add_parser("validate"); s.add_argument("spec")
    s = sub.add_parser("generate"); s.add_argument("spec"); s.add_argument("--rows", type=int, default=50)
    s.add_argument("--out", required=True); s.add_argument("--seed", type=int, default=7)
    s.add_argument("--xlsx-template"); s.add_argument("--hwpx-template"); s.add_argument("--hwpx-rows", type=int, default=3)
    s.add_argument(CONFIRM_FLAG, dest="confirm_fake", action="store_true", help=f"사용자가 「{CONFIRM_TEXT}」 를 확인했다는 표시(필수)")
    s = sub.add_parser("render"); s.add_argument("data")
    s.add_argument("--xlsx-template"); s.add_argument("--hwpx-template"); s.add_argument("--hwpx-rows", type=int, default=3)
    s = sub.add_parser("verify"); s.add_argument("spec"); s.add_argument("data")
    s = sub.add_parser("fill-text"); s.add_argument("data"); s.add_argument("texts")
    a = ap.parse_args(argv)
    try:
        if a.cmd == "presets":
            for p in list_presets():
                print(p)
            return 0
        if a.cmd == "show":
            print(f"# {PRESET_NOTICE}")
            print(resolve_spec_path(a.spec).read_text(encoding="utf-8")); return 0
        return {"validate": cmd_validate, "generate": cmd_generate, "render": cmd_render, "verify": cmd_verify, "fill-text": cmd_fill_text}[a.cmd](a)
    except SpecError as e:
        print(f"오류: {e}", file=sys.stderr); return 2


if __name__ == "__main__":
    sys.exit(main())

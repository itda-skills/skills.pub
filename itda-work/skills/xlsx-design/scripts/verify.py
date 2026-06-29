# -*- coding: utf-8 -*-
"""verify — XLSX 디자인 검증 게이트 (SPEC-OFFICE-DOC-GEN-DEEPEN-001 P5).

xlsx 의 한글 불변량: 한글이 담긴 셀의 폰트가 Korean-capable 이어야 한다(라틴 디스플레이
폰트로 박히면 의도와 다른 폴백 렌더 — 셀 단일 폰트라 docx 의 eastAsia 분리와 다른 형태).

층:
  (A) 콘텐츠 존재: 값이 든 셀이 하나도 없으면 빈 통합문서 [HARD]
  (B) 콘텐츠 대조: --tokens 의 필수 토큰이 셀 값에 존재 [HARD]
  (C) ★한글 셀 폰트: 한글 셀이 비-Korean-capable(라틴) 폰트 [HARD]
  (D) 구조 휴리스틱: 스타일 헤더(fill 셀) 0개·차트 0개 [advisory]
  (E) 렌더 빈 페이지 [advisory] (렌더 도구 부재 시 생략)

HARD GATE = (빈통합문서 + 토큰누락 + 한글_비안전폰트_셀) == 0 → PASS 시 exit 0.

사용:
  py -3 verify.py <xlsx> [--tokens tokens.txt] [--out DIR] [--dpi 130] [--no-render]
"""
from __future__ import annotations

import argparse
import json
import os
import sys

from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import render as render_mod  # noqa: E402
import sheetkit as sk  # noqa: E402


def _all_value_cells(wb):
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if c.value is not None and c.value != "":
                    yield ws, c


# ──────────────────────────────────────────────── (F) 색 대비(가독성)
def _wcag_lum(hex6):
    def _lin(x):
        x = x / 255.0
        return x / 12.92 if x <= 0.03928 else ((x + 0.055) / 1.055) ** 2.4
    r, g, b = int(hex6[0:2], 16), int(hex6[2:4], 16), int(hex6[4:6], 16)
    return 0.2126 * _lin(r) + 0.7152 * _lin(g) + 0.0722 * _lin(b)


def _wcag_contrast(c1, c2):
    l1, l2 = _wcag_lum(c1), _wcag_lum(c2)
    hi, lo = max(l1, l2), min(l1, l2)
    return (hi + 0.05) / (lo + 0.05)


def _argb_to_hex6(rgb):
    """openpyxl ARGB('FFRRGGBB') 또는 'RRGGBB' → 'RRGGBB'(없으면 None)."""
    if not isinstance(rgb, str):
        return None
    h = rgb.upper()
    if len(h) == 8:
        return h[2:]
    if len(h) == 6:
        return h
    return None


def _cell_fg(c):
    col = c.font.color if c.font is not None else None
    h = _argb_to_hex6(getattr(col, "rgb", None)) if col is not None else None
    return h or "000000"   # 기본 검정(흰 배경 위 고대비 — 거짓양성 회피)


def _cell_bg(c):
    fl = c.fill
    if fl is not None and getattr(fl, "fill_type", None) == "solid":
        h = _argb_to_hex6(getattr(fl.fgColor, "rgb", None))
        if h:
            return h
    return "FFFFFF"   # 기본 시트 = 흰색(샌드위치 본문)


def verify(xlsx_path, tokens=None, out_dir=None, dpi=130, do_render=True):
    xlsx_path = render_mod.winpath(xlsx_path)
    wb = load_workbook(xlsx_path)
    stem = os.path.splitext(os.path.basename(xlsx_path))[0]
    out_dir = out_dir or os.path.join(os.path.dirname(os.path.abspath(xlsx_path)), "_verify")
    os.makedirs(out_dir, exist_ok=True)

    issues = {
        "empty_workbook": [], "missing_content": [], "kr_unsafe_font": [],
        "low_contrast": [], "weak_contrast": [],
        "no_styled_header": [], "no_chart": [], "blank_page": [], "render_unavailable": [],
    }

    # (A)·(C)·(F) 전 셀 순회
    texts = []
    kr_total = 0
    fill_cells = 0
    for ws, c in _all_value_cells(wb):
        val = c.value
        texts.append(str(val))
        try:
            if c.fill is not None and c.fill.fill_type == "solid":
                fill_cells += 1
        except Exception:
            pass
        if sk.has_hangul(val):
            kr_total += 1
            fname = (c.font.name if c.font is not None else None) or ""
            if sk.is_latin_only_font(fname):
                issues["kr_unsafe_font"].append(
                    {"sheet": ws.title, "cell": c.coordinate,
                     "text": str(val)[:24], "font": fname})
        # (F) 셀 글자색 ↔ 채움색 대비. HARD<3.0 / ADVISORY 3.0~4.5.
        #   조건부서식(semantic_rules)의 up/down 색은 렌더 시 Excel 이 적용 — 정적 셀은
        #   ink(읽힘)를 유지하므로 여기선 거짓양성 없음(up/down 토큰도 보정됨).
        fg, bg = _cell_fg(c), _cell_bg(c)
        ratio = _wcag_contrast(fg, bg)
        rec = {"sheet": ws.title, "cell": c.coordinate, "text": str(val)[:24],
               "fg": "#" + fg, "bg": "#" + bg, "ratio": round(ratio, 2)}
        if ratio < 3.0:
            issues["low_contrast"].append(rec)
        elif ratio < 4.5:
            issues["weak_contrast"].append(rec)
    if not texts:
        issues["empty_workbook"].append({"note": "값이 든 셀이 없음"})

    # (B) 토큰 대조
    if tokens:
        compact = "".join(texts).replace(" ", "")
        for tok in tokens:
            if tok.replace(" ", "") not in compact:
                issues["missing_content"].append(tok)

    # (D) 구조 휴리스틱
    if texts and fill_cells == 0:
        issues["no_styled_header"].append({"note": "fill 셀 0개 — 스타일 헤더 부재 의심"})
    n_charts = sum(len(getattr(ws, "_charts", []) or []) for ws in wb.worksheets)
    if texts and n_charts == 0:
        issues["no_chart"].append({"note": "차트 0개(데이터만이면 정상일 수 있음)"})

    # (E) 렌더 + 빈 페이지
    pages = []
    if do_render:
        try:
            res = render_mod.render(xlsx_path, out_dir=os.path.join(out_dir, f"{stem}_render"), dpi=dpi)
            if res["engine"] is None:
                issues["render_unavailable"].append({"note": res.get("error")})
            pages = res.get("pages", [])
        except Exception as e:
            issues["render_unavailable"].append({"note": f"렌더 실패: {e}"})
    else:
        issues["render_unavailable"].append({"note": "do_render=False"})

    for pi, png in enumerate(pages, start=1):
        try:
            from PIL import Image
            im = Image.open(png).convert("RGB").resize((90, 90))
            cols = im.getcolors(90 * 90) or []
            if cols and max(cols, key=lambda x: x[0])[0] / (90 * 90) > 0.997:
                issues["blank_page"].append({"page": pi})
        except Exception:
            pass

    counts = {k: len(v) for k, v in issues.items()}
    gate = (counts["empty_workbook"] + counts["missing_content"]
            + counts["kr_unsafe_font"] + counts["low_contrast"])
    result = {
        "xlsx": xlsx_path, "hard_gate_pass": gate == 0, "counts": counts,
        "kr_cells_total": kr_total, "styled_fill_cells": fill_cells, "charts": n_charts,
        "render_pages": len(pages), "issues": issues,
    }
    json.dump(result, open(os.path.join(out_dir, f"{stem}.json"), "w", encoding="utf-8"),
              ensure_ascii=False, indent=2)
    return result


def main():
    for _stream in (sys.stdout, sys.stderr):
        try:
            _stream.reconfigure(encoding="utf-8")
        except Exception:
            pass
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    ap.add_argument("--tokens")
    ap.add_argument("--out")
    ap.add_argument("--dpi", type=int, default=130)
    ap.add_argument("--no-render", action="store_true")
    a = ap.parse_args()
    tokens = None
    if a.tokens and os.path.exists(a.tokens):
        tokens = [ln.strip() for ln in open(a.tokens, encoding="utf-8") if ln.strip()]
    r = verify(a.xlsx, tokens=tokens, out_dir=a.out, dpi=a.dpi, do_render=not a.no_render)
    c = r["counts"]
    print(f"[HARD] empty_wb={c['empty_workbook']} missing_tokens={c['missing_content']} "
          f"kr_unsafe_font={c['kr_unsafe_font']} low_contrast={c['low_contrast']} (한글 셀 {r['kr_cells_total']}개)")
    print(f"[ADVISORY] weak_contrast={c['weak_contrast']} no_styled_header={c['no_styled_header']} "
          f"no_chart={c['no_chart']} blank_page={c['blank_page']} render_unavailable={c['render_unavailable']} "
          f"(fill셀 {r['styled_fill_cells']}, 차트 {r['charts']}, 렌더 {r['render_pages']}p)")
    for o in r["issues"]["missing_content"]:
        print(f"  [토큰누락] '{o}'")
    for o in r["issues"]["kr_unsafe_font"]:
        print(f"  [HARD/한글] {o['sheet']}!{o['cell']} 비안전 폰트 '{o['font']}': '{o['text']}'")
    for o in r["issues"]["low_contrast"]:
        print(f"  [HARD/대비] {o['sheet']}!{o['cell']} {o['ratio']}:1 (fg {o['fg']} / bg {o['bg']}): '{o['text']}'")
    for o in r["issues"]["weak_contrast"][:8]:
        print(f"  [ADVISORY/대비] {o['sheet']}!{o['cell']} {o['ratio']}:1 (fg {o['fg']} / bg {o['bg']}): '{o['text']}'")
    for o in r["issues"]["render_unavailable"]:
        print(f"  [render] {o['note']}")
    print("HARD GATE:", "PASS" if r["hard_gate_pass"] else "FAIL")
    sys.exit(0 if r["hard_gate_pass"] else 1)


if __name__ == "__main__":
    main()

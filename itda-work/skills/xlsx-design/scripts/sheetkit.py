# -*- coding: utf-8 -*-
"""sheetkit — xlsx-design 공개 헬퍼 API (openpyxl 래퍼).

deckkit(pptx)·dockit(docx) 의 xlsx 형제. 디자인된 표·KPI·조건부서식·차트를
design-core 토큰으로 만든다.

★한글 정책(SPEC-OFFICE-DOC-GEN-DEEPEN-001 P5 — docx 와 다른 재설계):
  Excel 은 셀(또는 rich-text run) 폰트가 단일 이름이라 docx 처럼 ascii↔eastAsia
  분리 바인딩이 없다(분리는 테마 major/minor 에서만). 따라서 xlsx 의 한글 정책은
  "한글이 담긴 셀의 폰트를 Korean-capable 폰트(Malgun Gothic 우선)로 보장"하는 것.
  라틴/숫자 셀은 디스플레이 폰트, 한글 셀은 안전 고딕으로 셀 단위 분기한다.

백엔드: openpyxl(크로스플랫폼, Office 불필요)이 1급. 토큰은 design-core 의
`xlsx_styles()` 로부터 받는다(매체 중립 SSOT).
"""
from __future__ import annotations

import os
import platform

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ──────────────────────────────────────────── 색/폰트 유틸
def _hex6(value) -> str:
    h = str(value or "").strip().lstrip("#")
    if len(h) == 3:
        h = "".join(c * 2 for c in h)
    return h.upper()


def argb(value) -> str:
    """'#1E2761' → 'FF1E2761' (openpyxl 색 문자열, 불투명)."""
    return "FF" + _hex6(value)


def luminance(value) -> float:
    h = _hex6(value)
    return 0.299 * int(h[0:2], 16) + 0.587 * int(h[2:4], 16) + 0.114 * int(h[4:6], 16)


def on_color(value, dark="#1A1A1A", light="#FFFFFF") -> str:
    return light if luminance(value) < 140 else dark


def has_hangul(text) -> bool:
    return any("가" <= ch <= "힣" or "ㄱ" <= ch <= "ㅎ" for ch in str(text or ""))


# 한글 폰트 후보 → 파일 힌트(존재 탐지)
_KR_FILE_HINTS = {
    "Malgun Gothic": ["malgun.ttf"], "맑은 고딕": ["malgun.ttf"],
    "Noto Sans KR": ["NotoSansKR-Regular.otf", "NotoSansKR-Regular.ttf", "NotoSansCJKkr-Regular.otf"],
    "Pretendard": ["Pretendard-Regular.ttf", "Pretendard-Regular.otf", "PretendardVariable.ttf"],
    "나눔고딕": ["NanumGothic.ttf"], "NanumGothic": ["NanumGothic.ttf"],
}
_DEFAULT_KR_CANDIDATES = ["Malgun Gothic", "맑은 고딕", "Noto Sans KR", "Pretendard"]
# 라틴 전용으로 간주(한글 셀에 박히면 의심) — verify 트립와이어와 동일 어휘
_LATIN_ONLY = ("helvetica", "arial", "calibri", "times", "georgia", "garamond",
               "palatino", "impact", "courier", "verdana", "tahoma", "cambria",
               "didot", "futura", "roboto", "inter", "open sans", "lato")


def _font_dirs():
    sysname = platform.system()
    if sysname == "Windows":
        return [d for d in (os.path.join(os.environ.get("WINDIR", r"C:\Windows"), "Fonts"),
                            os.path.join(os.environ.get("LOCALAPPDATA", ""), "Microsoft", "Windows", "Fonts"))
                if d and os.path.isdir(d)]
    if sysname == "Darwin":
        return [d for d in ("/System/Library/Fonts", "/Library/Fonts",
                            os.path.expanduser("~/Library/Fonts")) if os.path.isdir(d)]
    return [d for d in ("/usr/share/fonts", "/usr/local/share/fonts",
                        os.path.expanduser("~/.fonts"), os.path.expanduser("~/.local/share/fonts"))
            if os.path.isdir(d)]


def _font_installed(name) -> bool:
    hints = _KR_FILE_HINTS.get(name)
    if not hints:
        return False
    for d in _font_dirs():
        for h in hints:
            if os.path.exists(os.path.join(d, h)):
                return True
    return False


def kr_font_name(candidates=None) -> str:
    """한글을 또렷하게 렌더하는 셀 폰트명(Korean-capable). 환경변수 XLSX_KR_FONT 우선."""
    env = os.environ.get("XLSX_KR_FONT")
    if env:
        return env
    for c in (candidates or _DEFAULT_KR_CANDIDATES):
        if _font_installed(c):
            return c
    return "Malgun Gothic" if platform.system() == "Windows" else (candidates or _DEFAULT_KR_CANDIDATES)[0]


def is_latin_only_font(name) -> bool:
    return bool(name) and any(k in name.strip().lower() for k in _LATIN_ONLY)


# ──────────────────────────────────────────── 워크북 / 디자인
def new_book() -> Workbook:
    wb = Workbook()
    return wb


def apply_design(wb, st: dict) -> dict:
    """xlsx_styles() dict 로 theme 구성(색·폰트) 후 wb 에 stash. 빌더가 기본값으로 읽음."""
    latin = st.get("latin_font") or "Calibri"
    kr = kr_font_name(st.get("korean_fallbacks"))
    theme = dict(st)
    theme["_latin"], theme["_kr"] = latin, kr
    # 통합문서 기본(Normal) 폰트를 한글 안전 폰트로(빈 셀 포함 일관)
    try:
        normal = wb._named_styles["Normal"]
        normal.font = Font(name=kr, size=float((st.get("type_scale") or {}).get("body", 10.5)),
                           color=argb(st.get("ink") or "#222222"))
    except Exception:
        pass
    wb._sheetkit_theme = theme  # type: ignore[attr-defined]
    return theme


def theme_of(wb) -> dict:
    return getattr(wb, "_sheetkit_theme", {}) or {}


def save_book(wb, path):
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    wb.save(path)
    return path


# ──────────────────────────────────────────── 셀/스타일 헬퍼
_HALIGN = {"left": "left", "center": "center", "right": "right"}


def _font_for(text, theme, latin=None, kr=None, guard=True):
    """셀 텍스트 언어에 맞는 폰트명 — 한글이면 kr, 아니면 라틴(가드)."""
    latin = latin or theme.get("_latin") or "Calibri"
    kr = kr or theme.get("_kr") or kr_font_name()
    if guard and has_hangul(text):
        return kr
    return latin


def thin_side(color):
    return Side(style="thin", color=argb(color))


def box_border(color):
    s = thin_side(color)
    return Border(left=s, right=s, top=s, bottom=s)


def set_cell(ws, coord, value, theme=None, latin=None, kr=None, size=None, bold=False,
             color=None, fill=None, align=None, valign="center", number_format=None,
             border_color=None, wrap=False, italic=False):
    """단일 셀 설정 + 한글 폰트 가드. coord 예: 'B2'."""
    theme = theme if theme is not None else {}
    c = ws[coord]
    c.value = value
    name = _font_for(value, theme, latin, kr)
    c.font = Font(name=name, size=size or 10.5, bold=bold, italic=italic,
                  color=argb(color) if color else (argb(theme.get("ink")) if theme.get("ink") else None))
    if fill:
        c.fill = PatternFill(fill_type="solid", fgColor=argb(fill))
    c.alignment = Alignment(horizontal=_HALIGN.get(align or "left"), vertical=valign, wrap_text=wrap)
    if number_format:
        c.number_format = number_format
    if border_color:
        c.border = box_border(border_color)
    return c


def styled_header(ws, row, headers, theme, start_col=1, fill=None, color=None,
                  border_color=None, align=None, size=10.5):
    """헤더 행: accent fill + 대비 글자색 + 테두리."""
    fill = fill or theme.get("accent") or theme.get("primary") or "#333333"
    color = color or on_color(fill)
    border_color = border_color or theme.get("border") or "#CCCCCC"
    aligns = align or (["left"] + ["right"] * (len(headers) - 1) if len(headers) > 1 else ["left"])
    for j, h in enumerate(headers):
        coord = f"{get_column_letter(start_col + j)}{row}"
        set_cell(ws, coord, h, theme=theme, bold=True, color=color, fill=fill,
                 align=(aligns[j] if j < len(aligns) else "left"), border_color=border_color, size=size)


def data_table(ws, start_row, headers, rows, theme, start_col=1, zebra=True,
               zebra_fill=None, header_fill=None, border_color=None, col_align=None,
               number_formats=None, value_colors=None, size=10):
    """디자인된 데이터 표(헤더 fill + zebra + 테두리 + 숫자서식 + 한글 폰트 가드).

    rows[i][j] = 값(스칼라). number_formats[j], value_colors[j]=hex(열별 글자색) 선택.
    반환 (first_data_row, last_row, first_col, last_col).
    """
    zebra_fill = zebra_fill or theme.get("surface") or "#F2F2F2"
    border_color = border_color or theme.get("border") or "#CCCCCC"
    ncol = len(headers)
    aligns = col_align or (["left"] + ["right"] * (ncol - 1) if ncol > 1 else ["left"])
    nfmt = number_formats or [None] * ncol
    vcol = value_colors or [None] * ncol

    styled_header(ws, start_row, headers, theme, start_col=start_col,
                  fill=header_fill, border_color=border_color, align=aligns)
    for i, rowvals in enumerate(rows):
        r = start_row + 1 + i
        band = zebra_fill if (zebra and i % 2 == 1) else None
        for j, val in enumerate(rowvals):
            coord = f"{get_column_letter(start_col + j)}{r}"
            set_cell(ws, coord, val, theme=theme,
                     align=(aligns[j] if j < len(aligns) else "left"),
                     fill=band, number_format=(nfmt[j] if j < len(nfmt) else None),
                     color=(vcol[j] if j < len(vcol) else None),
                     border_color=border_color, size=size)
    return (start_row + 1, start_row + len(rows), start_col, start_col + ncol - 1)


def kpi_block(ws, row, items, theme, start_col=1, value_color=None, label_color=None,
              value_size=18, label_size=9.5):
    """KPI 블록: [(value, label), ...] 를 값(큰)·라벨(작은) 2행으로."""
    # KPI 값은 라이트 시트 위 텍스트 — 저대비 브랜드색이면 design-core 가 보정한
    # primary_text 사용(대비 충족 프리셋은 primary 와 동일).
    value_color = value_color or theme.get("primary_text") or theme.get("primary") or theme.get("ink")
    label_color = label_color or theme.get("muted")
    for j, (value, label) in enumerate(items):
        col = get_column_letter(start_col + j)
        set_cell(ws, f"{col}{row}", value, theme=theme, bold=True, color=value_color,
                 size=value_size, align="left")
        set_cell(ws, f"{col}{row + 1}", label, theme=theme, color=label_color,
                 size=label_size, align="left")


def title_block(ws, runs, theme, row=1, span=1, start_col=1, size=18, color=None, fill=None):
    """제목 셀(필요 시 병합). runs: str 또는 (text) — 셀 단일 폰트라 단순 텍스트.

    글자색 자동: fill(밴드) 이 있으면 그 위 대비색(on_color), 없으면 라이트 시트 위
    가독 브랜드색(primary_text). 표지/제목 밴드를 다크 프리셋에서도 안전하게 만든다.
    """
    if color is None:
        color = (on_color(fill) if fill
                 else (theme.get("primary_text") or theme.get("primary") or theme.get("ink")))
    text = runs if isinstance(runs, str) else "".join(str(t) for t in runs)
    col0 = get_column_letter(start_col)
    if span > 1:
        ws.merge_cells(f"{col0}{row}:{get_column_letter(start_col + span - 1)}{row}")
    set_cell(ws, f"{col0}{row}", text, theme=theme, bold=True, size=size, color=color, fill=fill)
    return ws[f"{col0}{row}"]


def semantic_rules(ws, cell_range, theme, up=None, down=None):
    """조건부서식: 양수=up 색·음수=down 색(convention 반영 가능). 의미색 강조 자동화."""
    up = up or theme.get("up") or "#0E7C4A"
    down = down or theme.get("down") or "#C0392B"
    ws.conditional_formatting.add(
        cell_range, CellIsRule(operator="greaterThan", formula=["0"],
                               font=Font(color=argb(up), bold=True)))
    ws.conditional_formatting.add(
        cell_range, CellIsRule(operator="lessThan", formula=["0"],
                               font=Font(color=argb(down), bold=True)))


def add_bar_chart(ws, title, data_ref, cats_ref, theme, anchor, height=7.5, width=14,
                  palette=None, y_title=None):
    return _add_chart(BarChart(), ws, title, data_ref, cats_ref, theme, anchor,
                      height, width, palette, y_title)


def add_line_chart(ws, title, data_ref, cats_ref, theme, anchor, height=7.5, width=14,
                   palette=None, y_title=None):
    ch = LineChart()
    return _add_chart(ch, ws, title, data_ref, cats_ref, theme, anchor,
                      height, width, palette, y_title)


def _add_chart(chart, ws, title, data_ref, cats_ref, theme, anchor, height, width, palette, y_title):
    chart.title = title
    chart.height = height
    chart.width = width
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    pal = palette or theme.get("chart_palette") or [theme.get("primary"), theme.get("accent")]
    for i, s in enumerate(chart.series):
        try:
            s.graphicalProperties.solidFill = _hex6(pal[i % len(pal)])
            s.graphicalProperties.line.solidFill = _hex6(pal[i % len(pal)])
        except Exception:
            pass
    if y_title:
        chart.y_axis.title = y_title
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    ws.add_chart(chart, anchor)
    return chart


def set_columns(ws, widths, start_col=1):
    """열 폭 지정(문자 단위). widths: [12, 10, ...]."""
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(start_col + i)].width = w


def freeze(ws, coord):
    ws.freeze_panes = coord
